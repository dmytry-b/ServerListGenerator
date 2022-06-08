import ipaddress
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import shutil


USER = 'username'
PASSWORD = 'password'
SNMP_PORT = 161
SSH_PORT = 22
IPMI_PORT = 623
VMM_PORT = 8208
HTTPS_PORT = 443
TEMPLATE = 'Template.xlsx'
SMART_FILE = 'Smart.xlsx'
SCRIPT_FILE = ''

work_dir = os.getcwd()
filler = '#############################################'


def load_data():
    while True:
        try:
            s = input('Please enter the filename\n')
            file_name = os.path.join(work_dir, s)
            input_list = []
            # file_name = '0904.csv'
            ext_file_name = file_name.split('.')[-1]
            if ext_file_name == 'csv':
                input_list = load_from_csv(file_name)
            printf(f'The information about {len(input_list)} was added')
            break
        except FileNotFoundError:
            printf('WARNING! The file not found!', 1, 1)
    return input_list


def load_from_csv(file_path):
    f = open(file_path)
    input_csv_list_final = []
    try:
        input_csv_list = f.readlines()
    finally:
        f.close()
    for i in range(len(input_csv_list)):
        input_csv_list[i].replace(' ', '')
        input_csv_list[i].strip()
    for i in range(len(input_csv_list)):
        input_csv_list_final.append(input_csv_list[i].split(';'))
    for i in range(len(input_csv_list_final)):
        for j in range(len(input_csv_list_final[i])):
            input_csv_list_final[i][j].strip()
    return input_csv_list_final


def sort_ip(ip_list):
    wraped = True
    while wraped:
        wraped = False
        for i in range(len(ip_list) - 1):
            if ipaddress.ip_address(ip_list[i][0]) > ipaddress.ip_address(ip_list[i + 1][0]):
                ip_list[i], ip_list[i + 1] = ip_list[i + 1], ip_list[i]
                wraped = True
    return ip_list


def do_intervals(ip_list):
    interval_list = []
    ip = ip_list[0][0]
    for i in range(len(ip_list)):
        if i == len(ip_list) - 1:
            interval_list.append([ip, ip_list[i][0]])
        else:
            ip1 = ipaddress.ip_address(ip_list[i][0])
            ip2 = ipaddress.ip_address(ip_list[i + 1][0])
            ip1_tmp = ip_list[i][0].split('.')
            del ip1_tmp[3]
            subnet_ip1 = ipaddress.ip_network('.'.join(ip1_tmp) + '.0/24')
            if (ip1 == ip2 - 1) and (ip2 in subnet_ip1):
                continue
            else:
                interval_list.append([ip, ip_list[i][0]])
                ip = ip_list[i + 1][0]
    return interval_list


def fillin_smart_xlsx(intervals):
    template_path = os.path.join(work_dir, TEMPLATE)
    new_smart_xlsx_path = os.path.join(work_dir, SMART_FILE)
    shutil.copy2(template_path, new_smart_xlsx_path, follow_symlinks=True)
    wb = load_workbook(new_smart_xlsx_path)
    ws = wb.active

    for i in range(len(intervals)):
        row = i + 2
        ws.cell(row, 1).value = intervals[i][0]
        ws.cell(row, 2).value = intervals[i][1]
        ws.cell(row, 3).value = USER
        ws.cell(row, 4).value = PASSWORD
        ws.cell(row, 8).value = SNMP_PORT
        ws.cell(row, 9).value = SSH_PORT
        ws.cell(row, 10).value = IPMI_PORT
        ws.cell(row, 11).value = VMM_PORT
        ws.cell(row, 12).value = HTTPS_PORT

    wb.save(new_smart_xlsx_path)
    wb.close()


def printf(filled_string, filling=1, empty=0):
    for i in (range(empty)):
        print()
    for i in range(filling):
        print(filler)
    print(filled_string)
    for i in range(filling):
        print(filler)
    for i in (range(empty)):
        print()


def runner():
    ip_list = load_data()
    print(ip_list)
    print()
    ip_list = sort_ip(ip_list)
    print(ip_list)
    print()
    intervals = do_intervals(ip_list)
    print(intervals)
    fillin_smart_xlsx(intervals)


runner()
